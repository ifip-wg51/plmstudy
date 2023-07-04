import requests
import sqlite3
import openpyxl

class ICPLMReader:

    def __init__(self, dbFile, dbFrom, dbTo, isbns, plm23File):
        self.dbFile = dbFile
        self.dbFrom = dbFrom
        self.dbTo = dbTo
        self.isbns = isbns
        self.plm23File = plm23File

    def getPLMStudyDatabase(self):
        pubs=[]
        con = sqlite3.connect(self.dbFile)
        cur = con.cursor()
        #select all publications where source = "International Conference on Product Lifecycle Management"
        #Returns values from 2005 until 2015

        cur.execute('SELECT * FROM publications WHERE pubSource = "International Conference on Product Lifecycle Management"')
        sqlPubs = cur.fetchall()
        for sqlPub in sqlPubs:
            curKeys = con.cursor()
            keys =[]
            for sqlKey in cur.execute('SELECT tags. tagName FROM publicationTags INNER JOIN tags ON publicationTags.tagID=tags.tagID WHERE publicationTags.pubID=' + str(sqlPub[0])):
                keys.append(sqlKey[0])
            pubs.append({
                    'title': sqlPub[1],
                    'keywords': keys,
                    'year':  str(sqlPub[2])
                })
        con.close()
        return pubs

    def getSpringerAPI(self, isbn):
        url = 'http://api.springernature.com/meta/v2/json?q=isbn:' + isbn + '&p=100&api_key=93347ea55743e63a222615f3af8dce2f'
        response = requests.get(url)
        data = response.json()
        pubs=[]
        for pub in data['records']:
            if 'keyword' in pub.keys():
                publication = {
                    'title': pub['title'],
                    'keywords': pub['keyword'],
                    'year':  str(pub['publicationDate'][0:4])
                }
                #correction for 2021 (officially these papers are published in 2022, but we want to assign them to 2021)
                if isbn == '978-3-030-94335-6' or isbn == '978-3-030-94399-8':
                    publication['year'] = '2021'
                if isbn == '978-3-031-25182-5':
                    publication['year'] = '2022'
                if isbn == '978-3-030-42250-9':
                    publication['year'] = '2019'
                if isbn == '978-3-319-54660-5':
                    publication['year'] = '2016'
                pubs.append(publication)
            else:
                print("    !!! Publication without keyword: " + pub['title'])
        return pubs
    
    def getPLM23(self):
        pubs=[]
        workbook = openpyxl.load_workbook(self.plm23File)
        sheet = workbook.active
        
        for i in range(2, sheet.max_row + 1):
            keys = sheet.cell(row = i, column = 4).value.split(',')
            for i, key in enumerate(keys):
                keys[i] = key.strip()
            pubs.append({
                    'title': sheet.cell(row = i, column = 3).value,
                    'keywords': keys,
                    'year':  '2023'
                })
        return pubs



    def read(self):
        publications = []
        print("    Start reading IC PLM input Data")
        print("    ... get content from Springer API")
        for isbn in self.isbns:  
           pubs = self.getSpringerAPI(isbn)
           print('isbn: ' + isbn + ' count: ' + str(len(pubs)))
           publications.extend(pubs)
        print("    ... get content from PLM Study database")
        pubs = self.getPLMStudyDatabase()
        publications.extend(pubs)
        print("    ... get conteont from PLM23")
        pubs = self.getPLM23()
        publications.extend(pubs)
        print("    Done reading")
        return publications

